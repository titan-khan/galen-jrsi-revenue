#!/usr/bin/env python3
"""
JR Pilot — End-to-end loader (UPDATED hybrid + 3-files version).

Process all 3 input files dan upload ke Supabase:
  Phase 1: JR_Data_Dictionary.xlsx → kb.reference_docs (pgvector embeddings)
  Phase 2: Transaksi_2025.csv → aggregate avg PKB → upsert dim_jenken
  Phase 3: Palangka_Raya.csv → cleanse + classify + denormalize treatment → upsert registry_enriched
  Phase 4: (Optional) Transaksi_2025.csv → upsert transaksi_fact (raw)
  Phase 5: Verification (compare distribution vs framework v1.4 Sheet 2)

Key features:
  - Rule-based classification (7-segment per framework v1.4)
  - Treatment recommendation lookup dari treatment_v1.4.yaml
  - Idempotent (ON CONFLICT upsert)
  - Transactional safety (all-or-nothing per phase)
  - Output parquet untuk audit sebelum upload
  - Automatic verification dengan expected counts dari Sheet 2

Usage:
  python 11_load_pilot.py \\
    --dictionary inputs/JR_Data_Dictionary.xlsx \\
    --transaksi  inputs/Transaksi_2025.csv \\
    --registry   inputs/Palangka_Raya.csv \\
    --treatment-config treatment_v1.4.yaml \\
    --output-dir outputs/ \\
    --reference-date 2025-05-01

Environment variables (.env):
  SUPABASE_DB_HOST, SUPABASE_DB_USER, SUPABASE_DB_PASSWORD
  SUPABASE_DB_PORT (default 6543)
  SUPABASE_DB_NAME (default postgres)
  OPENAI_API_KEY (for Phase 1 embedding)

Reference date: 2025-05-01 (per framework v1.4)
"""
import argparse
import hashlib
import json
import os
import sys
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import psycopg2
import yaml
from psycopg2.extras import execute_values

# Optional: for embedding (Phase 1)
try:
    from openai import OpenAI
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False

# =============================================================================
# CONFIG
# =============================================================================

REFERENCE_DATE = date(2025, 5, 1)

EXPECTED_DISTRIBUTION = {
    # From framework v1.4 Sheet 2 (Matriks Segmen × Wilayah)
    "H1": 107967,
    "K1": 33789,
    "O1": 32916,
    "M1": 25039,
    "M2": 140966,
    "S1": 12756,
    "S2": 74544,
}
EXPECTED_TOTAL = sum(EXPECTED_DISTRIBUTION.values())  # 427,977

# =============================================================================
# DATABASE
# =============================================================================

def get_conn():
    """Connect to Supabase Postgres via pooler."""
    return psycopg2.connect(
        host=os.environ["SUPABASE_DB_HOST"],
        port=int(os.environ.get("SUPABASE_DB_PORT", "6543")),
        user=os.environ["SUPABASE_DB_USER"],
        password=os.environ["SUPABASE_DB_PASSWORD"],
        dbname=os.environ.get("SUPABASE_DB_NAME", "postgres"),
        sslmode="require",
    )


def insert_batch_manifest(conn, source_file, source_period, raw_md5, raw_count) -> str:
    batch_id = str(uuid.uuid4())
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO public.batch_manifest
              (batch_id, source_file_name, source_period, raw_md5_checksum,
               raw_row_count, status)
            VALUES (%s, %s, %s, %s, %s, 'uploading')
            """,
            (batch_id, source_file, source_period, raw_md5, raw_count),
        )
    return batch_id


def update_batch_manifest(conn, batch_id, status, loaded_count=None, notes=None):
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE public.batch_manifest
            SET status = %s,
                loaded_row_count = COALESCE(%s, loaded_row_count),
                notes = COALESCE(%s, notes)
            WHERE batch_id = %s
            """,
            (status, loaded_count, notes, batch_id),
        )

# =============================================================================
# PHASE 1 — DATA DICTIONARY EMBEDDING
# =============================================================================

def phase1_embed_dictionary(dictionary_path: Path, conn) -> int:
    """
    Read JR_Data_Dictionary.xlsx, chunk per row, embed via OpenAI,
    insert ke kb.reference_docs.

    Returns: number of chunks embedded.
    """
    print("\n[Phase 1/5] Reading & embedding data dictionary...")

    if not HAS_OPENAI:
        print("  ⚠ OpenAI not installed. Skip embedding (run `pip install openai` to enable).")
        return 0

    if not os.environ.get("OPENAI_API_KEY"):
        print("  ⚠ OPENAI_API_KEY not set. Skip embedding.")
        return 0

    # Read both sheets
    import openpyxl
    wb = openpyxl.load_workbook(dictionary_path, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header & section title
        section_title = None
        header_idx = None
        for idx, row in enumerate(rows):
            if row and row[0]:
                if section_title is None:
                    section_title = str(row[0])
                if str(row[0]) == "Nama Field":
                    header_idx = idx
                    break

        if header_idx is None:
            continue

        # Embed each field row as separate chunk
        for row in rows[header_idx + 1 :]:
            if not row[0]:
                continue
            field_name = str(row[0])
            tipe_data = str(row[1]) if row[1] else ""
            panjang = str(row[2]) if row[2] else ""
            keterangan = str(row[3]) if len(row) > 3 and row[3] else ""

            chunk_text = (
                f"[Section: {section_title}] "
                f"Field: {field_name} | "
                f"Type: {tipe_data} ({panjang}) | "
                f"Keterangan: {keterangan}"
            )
            chunks.append({
                "source": "data_dictionary",
                "chunk_text": chunk_text,
                "metadata": {
                    "section": section_title,
                    "field": field_name,
                    "type": tipe_data,
                    "length": panjang,
                },
            })

    print(f"  → {len(chunks)} chunks extracted dari dictionary")

    # Embed via OpenAI
    client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])
    print("  → Embedding via OpenAI text-embedding-3-small...")

    rows = []
    for i, chunk in enumerate(chunks):
        try:
            resp = client.embeddings.create(
                model="text-embedding-3-small",
                input=chunk["chunk_text"],
            )
            embedding = resp.data[0].embedding
            rows.append((
                chunk["source"],
                chunk["chunk_text"],
                json.dumps(chunk["metadata"]),
                embedding,
            ))
            if (i + 1) % 10 == 0:
                print(f"     {i + 1}/{len(chunks)} embedded")
            time.sleep(0.05)  # rate limit safety
        except Exception as e:
            print(f"     ✗ Failed chunk {i}: {e}")

    # Insert
    with conn.cursor() as cur:
        execute_values(
            cur,
            """INSERT INTO kb.reference_docs (source, chunk_text, chunk_metadata, embedding)
               VALUES %s""",
            rows,
            template="(%s, %s, %s::jsonb, %s::vector)",
        )
    conn.commit()
    print(f"  ✓ {len(rows)} chunks inserted ke kb.reference_docs")
    return len(rows)


# =============================================================================
# PHASE 2 — TRANSAKSI AGGREGATE → DIM_JENKEN
# =============================================================================

def phase2_aggregate_transaksi(transaksi_path: Path, conn, output_dir: Path) -> dict:
    """
    Read Transaksi_2025.csv, compute avg PKB per kode_jenken,
    upsert ke gold.dim_jenken.

    Returns: dict {kode_jenken: avg_pkb} untuk dipakai di Phase 3.
    """
    print("\n[Phase 2/5] Computing transaksi aggregates...")
    print(f"  → Reading {transaksi_path.name} (this is ~270MB)...")

    df = pd.read_csv(transaksi_path, dtype=str, keep_default_na=False)
    print(f"  → {len(df):,d} transaksi rows")

    # Cleanse minimum
    df["pokok_pkb"] = pd.to_numeric(df["pokok_pkb"], errors="coerce").fillna(0)
    df["kode_jenken"] = df["kode_jenken"].astype(str).str.strip()
    df["jenis_kendaraan"] = df["jenis_kendaraan"].astype(str).str.strip()

    # Filter only positive amounts (exclude is_correction edge cases)
    df_clean = df[df["pokok_pkb"] > 0].copy()
    print(f"  → {len(df_clean):,d} transaksi dengan pokok_pkb > 0")

    # Aggregate avg PKB per kode_jenken
    agg = df_clean.groupby("kode_jenken").agg(
        avg_pkb=("pokok_pkb", "mean"),
        n_transaksi=("pokok_pkb", "count"),
        primary_label=("jenis_kendaraan",
                        lambda s: s.mode().iloc[0] if len(s.mode()) > 0 else ""),
    ).reset_index()

    # Determine is_motor (R = Sepeda Motor)
    agg["is_motor"] = agg["kode_jenken"] == "R"

    # Save audit output
    output_dir.mkdir(parents=True, exist_ok=True)
    agg_path = output_dir / "transaksi_aggregated.parquet"
    agg.to_parquet(agg_path)
    print(f"  → Saved aggregate ke {agg_path}")

    print(f"\n  Avg PKB per kode_jenken:")
    print(f"  {'Kode':<6} {'Jenis':<25} {'Avg PKB':>15} {'N Trx':>10}")
    for _, row in agg.iterrows():
        print(f"  {row['kode_jenken']:<6} {row['primary_label']:<25} "
              f"{int(row['avg_pkb']):>15,d} {int(row['n_transaksi']):>10,d}")

    # Upsert ke dim_jenken
    print(f"\n  → Upserting {len(agg)} rows ke gold.dim_jenken...")
    rows = [
        (
            row["kode_jenken"],
            row["primary_label"],
            bool(row["is_motor"]),
            int(row["avg_pkb"]),
        )
        for _, row in agg.iterrows()
    ]

    with conn.cursor() as cur:
        execute_values(
            cur,
            """INSERT INTO gold.dim_jenken
                 (kode_jenken, jenis_kendaraan, is_motor, est_pkb_per_kendaraan)
               VALUES %s
               ON CONFLICT (kode_jenken) DO UPDATE SET
                 jenis_kendaraan = EXCLUDED.jenis_kendaraan,
                 is_motor = EXCLUDED.is_motor,
                 est_pkb_per_kendaraan = EXCLUDED.est_pkb_per_kendaraan""",
            rows,
        )
    conn.commit()
    print(f"  ✓ dim_jenken populated")

    # Return mapping untuk Phase 3
    return dict(zip(agg["kode_jenken"], agg["avg_pkb"].astype(int)))


# =============================================================================
# PHASE 3 — REGISTRY: CLEANSE + CLASSIFY + ENRICH + UPLOAD
# =============================================================================

def cleanse_registry(df: pd.DataFrame) -> pd.DataFrame:
    """Cleanse Palangka_Raya.csv per Data Assessment findings."""
    print(f"  → Cleansing {len(df):,d} rows...")

    # Trim string columns
    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype(str).str.strip()
            df.loc[df[c] == "", c] = None

    # Date parsing (YYYY-MM-DD format in registry)
    df["sd_notice"] = pd.to_datetime(
        df["sd_notice"], format="%Y-%m-%d", errors="coerce"
    ).dt.date
    df["tanggal_transaksi"] = pd.to_datetime(
        df["tanggal_transaksi"], format="%Y-%m-%d", errors="coerce"
    ).dt.date

    # Numerics
    df["thn_buat"] = pd.to_numeric(df["thn_buat"], errors="coerce").astype("Int32")

    # Filter SAMSAT PALANGKARAYA only (per Data Assessment finding:
    # the file has two-source merge, second source has different schema)
    n_before = len(df)
    df = df[df["nama_upt"] == "SAMSAT PALANGKARAYA"].copy()
    print(f"     Filtered SAMSAT PALANGKARAYA only: {n_before:,d} → {len(df):,d}")

    return df


def derive_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Compute derived columns needed for classification."""
    print(f"  → Deriving columns (durasi, has_history, usia, has_phone)...")

    df["has_payment_history"] = df["tanggal_transaksi"].notna()
    df["usia_kendaraan"] = (REFERENCE_DATE.year - df["thn_buat"]).astype("Int32")
    df["durasi_tunggakan_days"] = df["sd_notice"].apply(
        lambda d: (REFERENCE_DATE - d).days if pd.notna(d) else None
    ).astype("Int32")
    df["has_phone"] = (
        df["no_hp_masked"].notna()
        & (df["no_hp_masked"].astype(str).str.strip() != "")
    )
    return df


def classify_segment(row) -> str:
    """
    7-segment classification per framework v1.4.

    Priority order:
      1. Out-of-pyramid (S1, S2): never transacted
      2. In-pyramid: H1 → K1 → O1 → M1 → M2

    Edge case: durasi >5 thn dengan usia ≥20 thn reklasifikasi ke S2.
    """
    sd_days = row["durasi_tunggakan_days"]
    has_hist = row["has_payment_history"]
    usia = row["usia_kendaraan"]

    # Out-of-pyramid
    if not has_hist:
        if pd.notna(usia) and usia <= 15:
            return "S1"
        elif pd.notna(usia) and usia > 15:
            return "S2"
        else:
            return "unclassified"

    # In-pyramid: priority order
    if pd.isna(sd_days) or sd_days <= 0:
        return "H1"

    if 1 <= sd_days <= 90:
        return "K1"

    if 91 <= sd_days <= 365:
        return "O1"

    if 366 <= sd_days <= 730:
        return "M1"

    if 731 <= sd_days <= 1825:
        return "M2"

    if sd_days > 1825:
        if pd.notna(usia) and usia < 20:
            return "M2"
        else:
            return "S2"  # reklasifikasi: tunggakan kronis + tua = ghost

    return "unclassified"


def enrich_treatment(df: pd.DataFrame, treatment_config: dict) -> pd.DataFrame:
    """
    Lookup treatment per segmen → denormalize ke 4 kolom di setiap row.
    Source of truth: treatment_v1.4.yaml.
    """
    print(f"  → Enriching treatment columns dari treatment_v1.4.yaml...")

    segments_cfg = treatment_config["segments"]

    # Build flat lookup dicts
    kanal = {k: v.get("treatment_kanal_utama", "").strip() for k, v in segments_cfg.items()}
    amnesti = {k: v.get("treatment_kebijakan_amnesti", "").strip() for k, v in segments_cfg.items()}
    aksi = {k: v.get("treatment_aksi_utama", "").strip() for k, v in segments_cfg.items()}
    konv = {k: v.get("treatment_perkiraan_konversi", "").strip() for k, v in segments_cfg.items()}

    df["treatment_kanal_utama"] = df["segmen_kepatuhan"].map(kanal).fillna("")
    df["treatment_kebijakan_amnesti"] = df["segmen_kepatuhan"].map(amnesti).fillna("")
    df["treatment_aksi_utama"] = df["segmen_kepatuhan"].map(aksi).fillna("")
    df["treatment_perkiraan_konversi"] = df["segmen_kepatuhan"].map(konv).fillna("")

    return df


def enrich_est_pkb(df: pd.DataFrame, avg_pkb_map: dict) -> pd.DataFrame:
    """Map est_pkb_per_kendaraan dari avg PKB Phase 2."""
    print(f"  → Mapping est_pkb_per_kendaraan dari dim_jenken...")
    df["est_pkb_per_kendaraan"] = df["kode_jenis_kendaraan"].map(avg_pkb_map).fillna(0).astype(int)
    return df


def phase3_load_registry(
    registry_path: Path,
    treatment_config: dict,
    avg_pkb_map: dict,
    conn,
    output_dir: Path,
) -> tuple[str, dict]:
    """
    Full pipeline registry: cleanse → derive → classify → enrich → upload.

    Returns: (batch_id, distribution_dict)
    """
    print("\n[Phase 3/5] Loading registry (cleanse + classify + enrich + upload)...")

    # Read & checksum
    raw_bytes = registry_path.read_bytes()
    raw_md5 = hashlib.md5(raw_bytes).hexdigest()
    df = pd.read_csv(registry_path, dtype=str, keep_default_na=False)
    print(f"  → {len(df):,d} raw rows from {registry_path.name} (md5={raw_md5[:12]}...)")

    # Phase 3.1: Cleanse
    df = cleanse_registry(df)

    # Phase 3.2: Derive computed columns
    df = derive_columns(df)

    # Phase 3.3: Classify segments (THE rule engine)
    print(f"  → Classifying 7-segment per framework v1.4...")
    df["segmen_kepatuhan"] = df.apply(classify_segment, axis=1)

    distribution = df["segmen_kepatuhan"].value_counts().to_dict()
    print(f"\n  Segment distribution:")
    print(f"  {'Segmen':<8} {'Actual':>10} {'Expected':>10} {'Diff %':>10} ✓")
    for seg in ["H1", "K1", "O1", "M1", "M2", "S1", "S2", "unclassified"]:
        actual = distribution.get(seg, 0)
        expected = EXPECTED_DISTRIBUTION.get(seg, 0)
        diff_pct = (actual - expected) / expected * 100 if expected else 0
        marker = "✓" if abs(diff_pct) < 5 else ("⚠" if abs(diff_pct) < 10 else "✗")
        print(f"  {seg:<8} {actual:>10,d} {expected:>10,d} {diff_pct:>+9.2f}% {marker}")

    # Phase 3.4: Enrich treatment dari YAML config
    df = enrich_treatment(df, treatment_config)

    # Phase 3.5: Enrich est_pkb dari avg_pkb_map (Phase 2 output)
    df = enrich_est_pkb(df, avg_pkb_map)

    # Save audit output
    output_dir.mkdir(parents=True, exist_ok=True)
    enriched_path = output_dir / "enriched_registry.parquet"
    df.to_parquet(enriched_path)
    print(f"\n  → Saved enriched data ke {enriched_path}")

    # Phase 3.6: Insert batch manifest + bulk upload
    print(f"\n  → Inserting batch manifest...")
    batch_id = insert_batch_manifest(
        conn, registry_path.name, "2025", raw_md5, len(df)
    )
    conn.commit()
    print(f"     batch_id = {batch_id}")

    # Map kabupaten_id (PR only for pilot)
    df["kabupaten_id"] = 6271

    # Prepare records — match registry_enriched schema column order
    print(f"  → Bulk uploading {len(df):,d} rows ke gold.registry_enriched...")
    records = df[[
        "id",                         # vehicle_id
        "nopol_masked",
        "kabupaten_id",
        "kode_jenis_kendaraan",       # → kode_jenken
        "sd_notice",
        "tanggal_transaksi",
        "thn_buat",
        "no_hp_masked",
        "merek_kendaraan",
        "tipe",
        "bahan_bakar",
        "warna_plat",
        "kecamatan",
        "durasi_tunggakan_days",
        "has_payment_history",
        "usia_kendaraan",
        "has_phone",
        "segmen_kepatuhan",
        "treatment_kanal_utama",
        "treatment_kebijakan_amnesti",
        "treatment_aksi_utama",
        "treatment_perkiraan_konversi",
        "est_pkb_per_kendaraan",
    ]].copy()
    records["kelurahan"] = None
    records["batch_id"] = batch_id
    records["loaded_at"] = datetime.utcnow()
    records["source_period"] = "2025"

    # Convert to tuples (handle NaN → None)
    records = records.astype(object).where(records.notna(), None)
    tuples = [tuple(r) for r in records.itertuples(index=False, name=None)]

    with conn.cursor() as cur:
        execute_values(
            cur,
            """INSERT INTO gold.registry_enriched (
                vehicle_id, nopol_masked, kabupaten_id, kode_jenken,
                sd_notice, tanggal_transaksi, thn_buat, no_hp_masked,
                merek_kendaraan, tipe, bahan_bakar, warna_plat, kecamatan,
                durasi_tunggakan_days, has_payment_history, usia_kendaraan, has_phone,
                segmen_kepatuhan,
                treatment_kanal_utama, treatment_kebijakan_amnesti,
                treatment_aksi_utama, treatment_perkiraan_konversi,
                est_pkb_per_kendaraan,
                kelurahan, batch_id, loaded_at, source_period
              ) VALUES %s
              ON CONFLICT (vehicle_id) DO UPDATE SET
                segmen_kepatuhan = EXCLUDED.segmen_kepatuhan,
                treatment_kanal_utama = EXCLUDED.treatment_kanal_utama,
                treatment_kebijakan_amnesti = EXCLUDED.treatment_kebijakan_amnesti,
                treatment_aksi_utama = EXCLUDED.treatment_aksi_utama,
                treatment_perkiraan_konversi = EXCLUDED.treatment_perkiraan_konversi,
                est_pkb_per_kendaraan = EXCLUDED.est_pkb_per_kendaraan,
                durasi_tunggakan_days = EXCLUDED.durasi_tunggakan_days,
                has_payment_history = EXCLUDED.has_payment_history,
                usia_kendaraan = EXCLUDED.usia_kendaraan,
                has_phone = EXCLUDED.has_phone,
                loaded_at = EXCLUDED.loaded_at,
                batch_id = EXCLUDED.batch_id""",
            tuples,
            page_size=2000,
        )
    update_batch_manifest(conn, batch_id, "loaded", loaded_count=len(df))
    conn.commit()
    print(f"  ✓ {len(df):,d} records loaded ke registry_enriched")

    return batch_id, distribution


# =============================================================================
# PHASE 4 — OPTIONAL: RAW TRANSAKSI LOAD
# =============================================================================

def phase4_load_transaksi_raw(transaksi_path: Path, conn, skip: bool = False):
    """
    Load raw Transaksi_2025.csv ke gold.transaksi_fact.
    Optional — kalau pilot focus segmentasi only, bisa skip.
    """
    if skip:
        print("\n[Phase 4/5] Skipped (--skip-transaksi-raw)")
        return 0

    print("\n[Phase 4/5] Loading raw transaksi to gold.transaksi_fact...")
    print(f"  → Reading {transaksi_path.name} (~270MB, may take 1-2 minutes)...")
    df = pd.read_csv(transaksi_path, dtype=str, keep_default_na=False)
    n = len(df)

    # Cleanse
    df["paid_on"] = pd.to_datetime(
        df["paid_on"], format="%d/%m/%Y %H:%M", errors="coerce"
    )
    for col in ["pokok_pkb", "tunggakan_pokok_pkb", "pokok_bbnkb",
                "pokok_swdkllj", "tunggakan_pokok_swdkllj",
                "denda_swdkllj", "tunggakan_denda_swdkllj"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["total_amount"] = (
        df["pokok_pkb"] + df["tunggakan_pokok_pkb"] +
        df["pokok_bbnkb"] + df["pokok_swdkllj"] +
        df["tunggakan_pokok_swdkllj"] + df["denda_swdkllj"] +
        df["tunggakan_denda_swdkllj"]
    )

    # Map kabupaten/upt (skip for now — only what's in transaksi)
    df["kabupaten_id"] = pd.to_numeric(df["kabupaten_id"], errors="coerce").astype("Int32")
    df["upt_id"] = pd.to_numeric(df["upt_id"], errors="coerce").astype("Int32")

    print(f"  → Cleansed {n:,d} rows. Total amount: Rp {df['total_amount'].sum():,.0f}")

    # Bulk insert (simple, no upsert — assume re-load drops table first)
    records = df[[
        "id", "paid_on", "kabupaten_id", "upt_id", "kode_jenken",
        "id_layanan", "pokok_pkb", "tunggakan_pokok_pkb", "pokok_bbnkb",
        "pokok_swdkllj", "tunggakan_pokok_swdkllj",
        "denda_swdkllj", "tunggakan_denda_swdkllj", "total_amount",
    ]].copy()
    records["nomor_polisi_masked"] = df["nomor_polisi_masked"]
    records = records.astype(object).where(records.notna(), None)
    tuples = [tuple(r) for r in records.itertuples(index=False, name=None)]

    print(f"  → Uploading {n:,d} rows...")
    with conn.cursor() as cur:
        # Truncate first (one-shot pilot semantics)
        cur.execute("TRUNCATE gold.transaksi_fact")
        execute_values(
            cur,
            """INSERT INTO gold.transaksi_fact (
                transaksi_id, paid_on, kabupaten_id, upt_id, kode_jenken,
                id_layanan,
                pokok_pkb, tunggakan_pokok_pkb, pokok_bbnkb,
                pokok_swdkllj, tunggakan_pokok_swdkllj,
                denda_swdkllj, tunggakan_denda_swdkllj, total_amount,
                vehicle_bucket
              ) VALUES %s""",
            tuples,
            page_size=5000,
        )
    conn.commit()
    print(f"  ✓ {n:,d} transaksi loaded")
    return n


# =============================================================================
# PHASE 5 — VERIFICATION + REFRESH MV
# =============================================================================

def phase5_verify(conn, distribution: dict, output_dir: Path):
    """Final verification + refresh materialized views."""
    print("\n[Phase 5/5] Verification + refresh materialized views...")

    # Refresh MVs
    print("  → Refreshing materialized views...")
    with conn.cursor() as cur:
        for mv in ["gold_plus.agg_segmen_kabupaten", "gold_plus.agg_segmen_jenken"]:
            try:
                cur.execute(f"REFRESH MATERIALIZED VIEW {mv}")
                print(f"     ✓ {mv}")
            except psycopg2.errors.UndefinedTable:
                print(f"     ⚠ {mv} not yet exists, skip")
                conn.rollback()
                continue
    conn.commit()

    # Verification queries
    print("\n  → Running verification queries...")
    with conn.cursor() as cur:
        # 1. Total rows
        cur.execute("SELECT COUNT(*) FROM gold.registry_enriched")
        total = cur.fetchone()[0]

        # 2. Distribution
        cur.execute("""
            SELECT segmen_kepatuhan, COUNT(*)
            FROM gold.registry_enriched GROUP BY 1 ORDER BY 1
        """)
        db_dist = dict(cur.fetchall())

        # 3. Phone availability per segmen
        cur.execute("""
            SELECT segmen_kepatuhan,
                   ROUND(AVG(CASE WHEN has_phone THEN 1.0 ELSE 0.0 END)::NUMERIC, 4)
            FROM gold.registry_enriched GROUP BY 1 ORDER BY 1
        """)
        phone_pct = dict(cur.fetchall())

        # 4. Treatment populated check
        cur.execute("""
            SELECT segmen_kepatuhan, COUNT(DISTINCT treatment_kanal_utama)
            FROM gold.registry_enriched GROUP BY 1 ORDER BY 1
        """)
        treatment_check = dict(cur.fetchall())

    # Build report
    report = []
    report.append("=" * 70)
    report.append("VERIFICATION REPORT")
    report.append("=" * 70)
    report.append(f"Generated: {datetime.now()}")
    report.append(f"Reference date: {REFERENCE_DATE}")
    report.append(f"Total rows in registry_enriched: {total:,d}")
    report.append(f"Expected total (Sheet 2): {EXPECTED_TOTAL:,d}")
    report.append(f"Diff: {total - EXPECTED_TOTAL:+,d}")
    report.append("")
    report.append(f"{'Segmen':<8} {'Actual':>10} {'Expected':>10} {'Diff%':>8} {'Phone%':>8} {'Treat':>6} ✓")
    all_pass = True
    for seg in ["H1", "K1", "O1", "M1", "M2", "S1", "S2"]:
        actual = db_dist.get(seg, 0)
        expected = EXPECTED_DISTRIBUTION[seg]
        diff = (actual - expected) / expected * 100 if expected else 0
        phone = phone_pct.get(seg, 0)
        n_treat = treatment_check.get(seg, 0)

        ok = abs(diff) < 5 and n_treat == 1
        if not ok:
            all_pass = False
        marker = "✓" if ok else "✗"
        report.append(f"  {seg:<8} {actual:>10,d} {expected:>10,d} {diff:>+7.2f}% "
                      f"{float(phone):>7.2%} {n_treat:>6d} {marker}")

    report.append("")
    report.append("OVERALL: " + ("✓ PASS" if all_pass else "✗ FAIL"))
    if not all_pass:
        report.append("Investigation needed: check enriched_registry.parquet untuk debug")

    report_text = "\n".join(report)
    print("\n" + report_text)

    # Save report
    report_path = output_dir / "verification_report.txt"
    report_path.write_text(report_text)
    print(f"\n  → Report saved ke {report_path}")

    return all_pass


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="JR Pilot loader — hybrid + 3-files approach",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--dictionary", required=True, type=Path)
    parser.add_argument("--transaksi", required=True, type=Path)
    parser.add_argument("--registry", required=True, type=Path)
    parser.add_argument("--treatment-config", required=True, type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("outputs"))
    parser.add_argument("--reference-date", default="2025-05-01")
    parser.add_argument("--skip-dictionary", action="store_true",
                        help="Skip Phase 1 (dictionary embedding)")
    parser.add_argument("--skip-transaksi-raw", action="store_true",
                        help="Skip Phase 4 (raw transaksi load)")
    args = parser.parse_args()

    # Override REFERENCE_DATE if specified
    global REFERENCE_DATE
    REFERENCE_DATE = datetime.strptime(args.reference_date, "%Y-%m-%d").date()

    # Validate inputs exist
    for f in [args.dictionary, args.transaksi, args.registry, args.treatment_config]:
        if not f.exists():
            print(f"ERROR: file not found: {f}")
            sys.exit(1)

    # Load treatment config
    print(f"Loading treatment config dari {args.treatment_config}...")
    with open(args.treatment_config) as f:
        treatment_config = yaml.safe_load(f)

    # Verify config
    expected_segments = treatment_config["verification"]["expected_segments"]
    config_segments = list(treatment_config["segments"].keys())
    missing = set(expected_segments) - set(config_segments)
    if missing:
        print(f"ERROR: treatment config missing segments: {missing}")
        sys.exit(1)
    print(f"✓ Treatment config valid (7 segmen)")

    # Connect
    print("\nConnecting to Supabase...")
    conn = get_conn()
    print("✓ Connected")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    try:
        # Phase 1: Dictionary embedding
        if not args.skip_dictionary:
            phase1_embed_dictionary(args.dictionary, conn)
        else:
            print("\n[Phase 1/5] Skipped (--skip-dictionary)")

        # Phase 2: Transaksi aggregate → dim_jenken
        avg_pkb_map = phase2_aggregate_transaksi(
            args.transaksi, conn, args.output_dir
        )

        # Phase 3: Registry classify + enrich + upload
        batch_id, distribution = phase3_load_registry(
            args.registry, treatment_config, avg_pkb_map, conn, args.output_dir
        )

        # Phase 4: Optional raw transaksi
        phase4_load_transaksi_raw(
            args.transaksi, conn, skip=args.skip_transaksi_raw
        )

        # Phase 5: Verification + MV refresh
        passed = phase5_verify(conn, distribution, args.output_dir)

        if passed:
            print("\n" + "=" * 70)
            print("✓ PILOT DATA LOADED SUCCESSFULLY")
            print("=" * 70)
            print(f"  batch_id: {batch_id}")
            print(f"  outputs: {args.output_dir}")
            print(f"  next steps: Day 5 (KB embedding), Day 6 (Edge Functions)")
        else:
            print("\n⚠ Verification FAILED. Check verification_report.txt for details.")
            sys.exit(2)

    except Exception as e:
        print(f"\n❌ Failed: {e}")
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
